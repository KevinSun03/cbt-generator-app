from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Sequence
import re
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


LA_TZ = ZoneInfo("America/Los_Angeles")


def today_la() -> date:
    return datetime.now(LA_TZ).date()


@dataclass
class AttendanceRow:
    name: str
    company: str
    time_in: str | None = None
    time_out: str | None = None
    note: str | None = None


def clean_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = s.replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_name(v) -> str:
    s = clean_text(v)
    s = s.strip(" -–—_/\\")
    return s


def is_bad_name(name: str) -> bool:
    if not name:
        return True

    low = name.lower().strip()

    exact_bad = {
        "name",
        "full name",
        "employee",
        "employee name",
        "姓名",
        "no",
        "编号",
    }

    if low in exact_bad:
        return True

    if "attendance" in low or "date" in low or "公司" in name:
        return True

    return False


def normalize_time(v) -> str | None:
    """Return Excel/Python/string times as H:MM, with no leading zero on hour."""
    if v is None:
        return None

    if isinstance(v, datetime):
        return f"{v.hour}:{v.minute:02d}"

    if isinstance(v, time):
        return f"{v.hour}:{v.minute:02d}"

    if isinstance(v, timedelta):
        total = int(v.total_seconds())
        total %= 24 * 3600
        h = total // 3600
        m = (total % 3600) // 60
        return f"{h}:{m:02d}"

    if isinstance(v, (int, float)):
        # Excel serial fraction for time.
        if 0 <= float(v) < 1:
            total = int(round(float(v) * 24 * 3600))
            total %= 24 * 3600
            h = total // 3600
            m = (total % 3600) // 60
            return f"{h}:{m:02d}"
        return str(v)

    s = clean_text(v)
    if not s:
        return None

    s = s.replace("；", ":").replace(";", ":").replace("：", ":").replace(".", ":")
    s = re.sub(r"\s+", "", s)
    s = s.lower().replace("am", "").replace("pm", "")

    m = re.search(r"(\d{1,2})(?::(\d{1,2}))?", s)
    if not m:
        return clean_text(v)

    h = int(m.group(1))
    minute = int(m.group(2) or 0)
    return f"{h}:{minute:02d}"


def extract_date_from_workbook(wb) -> date | None:
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            for v in row:
                if isinstance(v, datetime):
                    return v.date()

                if isinstance(v, date):
                    return v

                s = clean_text(v)
                if not s:
                    continue

                m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", s)
                if m:
                    month, day, year = map(int, m.groups())
                    if year < 100:
                        year += 2000

                    try:
                        return date(year, month, day)
                    except ValueError:
                        pass

    return None


def _header_blocks(ws) -> list[tuple[int, int, int, int, int]]:
    """Return (header_row, name_col, position_col, in_col, out_col)."""
    blocks = []
    max_scan_row = min(ws.max_row, 20)

    for r in range(1, max_scan_row + 1):
        for c in range(1, ws.max_column + 1):
            v = clean_text(ws.cell(r, c).value).lower()

            if v in {"full name", "name"} or "full name" in v:
                blocks.append((r, c, c + 1, c + 2, c + 3))

            elif clean_text(ws.cell(r, c).value) == "姓名":
                nearby = " ".join(
                    clean_text(ws.cell(r, cc).value)
                    for cc in range(c, min(ws.max_column, c + 4) + 1)
                )
                if "上班" in nearby or "下班" in nearby:
                    blocks.append((r, c, c + 1, c + 2, c + 3))

    seen = set()
    unique = []

    for item in blocks:
        key = (item[0], item[1])
        if key not in seen:
            seen.add(key)
            unique.append(item)

    return unique


def extract_attendance_sheet_rows(
    wb,
    company: str,
    default_note: str | None = None,
) -> list[AttendanceRow]:
    """Parse NOVA/Newstart daily sign-in sheets."""
    rows: list[AttendanceRow] = []

    for ws in wb.worksheets:
        blocks = _header_blocks(ws)

        if not blocks:
            continue

        for header_row, name_col, pos_col, in_col, out_col in blocks:
            for r in range(header_row + 1, ws.max_row + 1):
                name = normalize_name(ws.cell(r, name_col).value)

                if is_bad_name(name):
                    continue

                if "备注" in name or "note" in name.lower():
                    continue

                time_in = normalize_time(ws.cell(r, in_col).value)
                time_out = normalize_time(ws.cell(r, out_col).value)
                pos = clean_text(ws.cell(r, pos_col).value) or default_note

                if not time_in and not time_out and not pos:
                    continue

                rows.append(
                    AttendanceRow(
                        name=name,
                        company=company,
                        time_in=time_in,
                        time_out=time_out,
                        note=pos,
                    )
                )

    return rows


def parse_nova(wb) -> list[AttendanceRow]:
    return extract_attendance_sheet_rows(wb, "NOVA")


NEWSTART_NOTE_MAP = {
    "林虹": "LA-收发货管理",
    "孙亦可": "LA-收发货",
    "祁奕帆": "早班收发货",
    "祁亦帆": "早班收发货",
    "杨俊杰": "签到管理",
    "王竹龙": "晚班叉车",
    "杨先忠": "晚班叉车",
    "彭宇": "",
    "李少龙": "晚班叉车",
    "周斌": "早班叉车",
    "谢庆伟": "晚班叉车",
}


def _parse_newstart_sheet(ws, company: str = "Newstart") -> list[AttendanceRow]:
    rows: list[AttendanceRow] = []
    blocks = _header_blocks(ws)

    if not blocks:
        return rows

    current_section_note: str | None = None

    for header_row, name_col, pos_col, in_col, out_col in blocks:
        for r in range(header_row + 1, ws.max_row + 1):
            raw_name = clean_text(ws.cell(r, name_col).value)

            if not raw_name:
                continue

            if "备注" in raw_name or "note" in raw_name.lower():
                parts = re.split(r"[:：]", raw_name, maxsplit=1)

                if len(parts) == 2:
                    current_section_note = clean_text(parts[1])
                else:
                    current_section_note = raw_name.replace("备注", "").strip(":： ")

                continue

            name = normalize_name(raw_name)

            if is_bad_name(name):
                continue

            time_in = normalize_time(ws.cell(r, in_col).value)
            time_out = normalize_time(ws.cell(r, out_col).value)
            pos = clean_text(ws.cell(r, pos_col).value)

            if not time_in and not time_out and not pos:
                continue

            note = pos or current_section_note or NEWSTART_NOTE_MAP.get(name, "")

            rows.append(
                AttendanceRow(
                    name=name,
                    company=company,
                    time_in=time_in,
                    time_out=time_out,
                    note=note,
                )
            )

    return rows


def _newstart_sheet_is_tt(ws) -> bool:
    """Detect TT sheet even when sheet tab is not named TT."""
    title = clean_text(ws.title).upper().replace(" ", "")

    if title == "TT" or "TT" in title:
        return True

    max_row = min(ws.max_row, 15)
    max_col = min(ws.max_column, 15)

    for r in range(1, max_row + 1):
        row_values = [
            clean_text(ws.cell(r, c).value)
            for c in range(1, max_col + 1)
        ]

        row_joined = " ".join(row_values).upper().replace(" ", "")

        if ("公司" in row_joined or "COMPANY" in row_joined) and "TT" in row_joined:
            return True

        for c in range(1, max_col + 1):
            cell = clean_text(ws.cell(r, c).value).upper().replace(" ", "")

            if cell in {"公司", "COMPANY"}:
                nearby = [
                    clean_text(ws.cell(r, cc).value).upper().replace(" ", "")
                    for cc in range(c + 1, min(max_col, c + 5) + 1)
                ]

                if "TT" in nearby:
                    return True

    return False


def parse_newstart(wb, tt_only: bool = True) -> list[AttendanceRow]:
    rows: list[AttendanceRow] = []

    for ws in wb.worksheets:
        is_tt = _newstart_sheet_is_tt(ws)

        if tt_only and not is_tt:
            continue

        parsed = _parse_newstart_sheet(ws)

        if not tt_only and not is_tt:
            for row in parsed:
                if not row.note:
                    row.note = ws.title

        rows.extend(parsed)

    return rows


WEEKDAY_ALIASES = {
    0: {"monday", "mon", "lunes"},
    1: {"tuesday", "tue", "tues", "martes"},
    2: {"wednesday", "wed", "miercoles", "miércoles"},
    3: {"thursday", "thu", "thur", "thurs", "jueves"},
    4: {"friday", "fri", "viernes"},
    5: {"saturday", "sat", "sabado", "sábado"},
    6: {"sunday", "sun", "domingo"},
}


def _weekday_key(v) -> str:
    return clean_text(v).lower().strip(" .:-")


def _compact_text(v) -> str:
    """Lowercase text with spaces and punctuation removed."""
    s = clean_text(v).lower()
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", s)


def _hrn_shift_from_text(v) -> str | None:
    """Detect first/second shift from title or top worksheet text."""
    t = _compact_text(v)

    if not t:
        return None

    first_markers = {
        "firstshift",
        "1stshift",
        "shift1",
        "first",
        "1st",
        "第一班",
        "一班",
        "早班",
    }

    second_markers = {
        "secondshift",
        "2ndshift",
        "shift2",
        "second",
        "2nd",
        "第二班",
        "二班",
        "晚班",
    }

    has_first = any(marker in t for marker in first_markers)
    has_second = any(marker in t for marker in second_markers)

    if has_second and not has_first:
        return "second"

    if has_first and not has_second:
        return "first"

    return None


def _hrn_sheet_shift(ws) -> str | None:
    """Return 'first', 'second', or None."""
    title_shift = _hrn_shift_from_text(ws.title)

    if title_shift:
        return title_shift

    max_row = min(ws.max_row, 15)
    max_col = min(ws.max_column, 12)

    for r in range(1, max_row + 1):
        row_text = " ".join(
            clean_text(ws.cell(r, c).value)
            for c in range(1, max_col + 1)
        )

        row_shift = _hrn_shift_from_text(row_text)

        if row_shift:
            return row_shift

    return None


def _is_hrn_bad_employee_row(name: str) -> bool:
    if is_bad_name(name):
        return True

    low = name.lower().strip()

    bad_keywords = [
        "total",
        "subtotal",
        "hours",
        "signature",
        "supervisor",
        "manager",
        "question",
        "concerns",
        "please reach out",
        "si tiene",
    ]

    return any(keyword in low for keyword in bad_keywords)


def _find_hrn_weekly_blocks(ws) -> list[tuple[int, int, int, int]]:
    """Find HRN weekly attendance blocks.

    Returns:
        employee_header_row, employee_col, weekday_row, inout_row
    """
    blocks: list[tuple[int, int, int, int]] = []

    max_scan_row = min(ws.max_row, 60)
    max_scan_col = min(ws.max_column, 30)
    all_weekday_aliases = set().union(*WEEKDAY_ALIASES.values())

    for r in range(1, max_scan_row + 1):
        for c in range(1, max_scan_col + 1):
            cell_value = clean_text(ws.cell(r, c).value).lower()

            if cell_value not in {"employee", "employee name"}:
                continue

            weekday_row = r
            inout_row = r + 1

            nearby_weekdays = [
                _weekday_key(ws.cell(weekday_row, cc).value)
                for cc in range(c + 1, min(ws.max_column, c + 16) + 1)
            ]

            if any(day in all_weekday_aliases for day in nearby_weekdays):
                blocks.append((r, c, weekday_row, inout_row))

    return blocks


def _selected_weekday_pair(
    ws,
    weekday_row: int,
    inout_row: int,
    start_col: int,
    target_date: date | None,
) -> tuple[int, int] | None:
    if target_date is None:
        return None

    wanted_weekday_names = WEEKDAY_ALIASES[target_date.weekday()]

    for c in range(start_col, ws.max_column + 1):
        day_label = _weekday_key(ws.cell(weekday_row, c).value)

        if day_label not in wanted_weekday_names:
            continue

        in_col = c
        out_col = c + 1 if c + 1 <= ws.max_column else None

        if out_col is None:
            return None

        in_label = clean_text(ws.cell(inout_row, in_col).value).lower()
        out_label = clean_text(ws.cell(inout_row, out_col).value).lower()

        if in_label == "in" and out_label == "out":
            return in_col, out_col

        # HRN keeps each weekday as a two-column pair.
        return in_col, out_col

    return None


def _parse_hrn_weekly_sheet(
    ws,
    target_date: date | None = None,
) -> list[AttendanceRow]:
    rows: list[AttendanceRow] = []

    for header_row, employee_col, weekday_row, inout_row in _find_hrn_weekly_blocks(ws):
        selected_pair = _selected_weekday_pair(
            ws=ws,
            weekday_row=weekday_row,
            inout_row=inout_row,
            start_col=employee_col + 1,
            target_date=target_date,
        )

        if selected_pair is None:
            continue

        in_col, out_col = selected_pair

        for r in range(inout_row + 1, ws.max_row + 1):
            if clean_text(ws.cell(r, employee_col).value).lower() in {"employee", "employee name"}:
                break

            name = normalize_name(ws.cell(r, employee_col).value)

            if _is_hrn_bad_employee_row(name):
                continue

            time_in = normalize_time(ws.cell(r, in_col).value)
            time_out = normalize_time(ws.cell(r, out_col).value)

            if not time_in and not time_out:
                continue

            rows.append(
                AttendanceRow(
                    name=name,
                    company="HRN",
                    time_in=time_in,
                    time_out=time_out,
                    note="LA-TT分拣",
                )
            )

    return rows


def _select_hrn_second_shift_sheets(wb) -> list:
    """Select only HRN Second Shift sheets.

    Rules:
    1. If sheet is clearly Second Shift, use it.
    2. If multiple weekly sheets exist but names are unclear, use only the last weekly sheet.
    3. If only one weekly sheet exists, use it for backward compatibility.
    """
    weekly_sheets = [
        ws for ws in wb.worksheets
        if _find_hrn_weekly_blocks(ws)
    ]

    if not weekly_sheets:
        return []

    second_shift_sheets = [
        ws for ws in weekly_sheets
        if _hrn_sheet_shift(ws) == "second"
    ]

    if second_shift_sheets:
        return second_shift_sheets

    if len(weekly_sheets) >= 2:
        return [weekly_sheets[-1]]

    return weekly_sheets


def _parse_hrn_weekly(
    wb,
    target_date: date | None = None,
) -> list[AttendanceRow]:
    rows: list[AttendanceRow] = []

    second_shift_sheets = _select_hrn_second_shift_sheets(wb)

    for ws in second_shift_sheets:
        rows.extend(_parse_hrn_weekly_sheet(ws, target_date=target_date))

    return rows


def parse_hrn(wb, target_date: date | None = None) -> list[AttendanceRow]:
    """Parse HRN workbook.

    HRN rule:
        Weekly workbook = Second Shift only.
        Never combine First Shift and Second Shift.
    """
    has_weekly_format = any(_find_hrn_weekly_blocks(ws) for ws in wb.worksheets)

    if has_weekly_format:
        return _parse_hrn_weekly(wb, target_date=target_date)

    attendance_rows = extract_attendance_sheet_rows(
        wb,
        "HRN",
        default_note="LA-TT分拣",
    )

    for row in attendance_rows:
        if not row.note:
            row.note = "LA-TT分拣"

    return attendance_rows


def _format_date_display(d: date) -> str:
    return f"{d.month}/{d.day}/{d.year}"


def _mmdd(d: date) -> str:
    return f"{d.month:02d}{d.day:02d}"


def build_cbt_workbook(
    company_rows: Sequence[tuple[str, Sequence[AttendanceRow]]],
    work_date: date,
) -> Workbook:
    if not company_rows:
        raise ValueError("No company data was provided. Select/upload at least one spreadsheet.")

    wb = Workbook()
    wb.remove(wb.active)

    for company, rows in company_rows:
        ws = wb.create_sheet(f"{company}_{_mmdd(work_date)}")
        write_cbt_sheet(ws, rows, company, work_date)

    return wb


def write_cbt_sheet(
    ws,
    rows: Sequence[AttendanceRow],
    company: str,
    work_date: date,
):
    widths = {
        "A": 9,
        "B": 28,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 26,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 32

    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    ws["A1"] = "CBT考勤"
    ws["A2"] = (
        f"日期：{_format_date_display(work_date)}　　　　"
        f"公司：{company}　　　　"
        f"填表人：________"
    )

    headers = ["编号", "姓名", "劳务公司", "上班时间", "下班时间", "备注"]

    for col_idx, h in enumerate(headers, 1):
        ws.cell(3, col_idx).value = h

    blue = "2F75B5"
    light_blue = "BDD7EE"
    band_blue = "D9E8FA"
    band_white = "F3F6FC"
    border_color = "A6A6A6"

    title_fill = PatternFill("solid", fgColor=blue)
    meta_fill = PatternFill("solid", fgColor="DDEBF7")
    header_fill = PatternFill("solid", fgColor=light_blue)
    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", size=12)

    ws["A1"].fill = title_fill
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")

    ws["A2"].fill = meta_fill
    ws["A2"].font = Font(name="Arial", size=12, color="1F1F1F")

    for c in range(1, 7):
        ws.cell(3, c).fill = header_fill
        ws.cell(3, c).font = Font(name="Arial", size=12, bold=True, color="1F4E79")

    for idx, row in enumerate(rows, 1):
        excel_row = idx + 3

        values = [
            idx,
            row.name,
            row.company,
            row.time_in or "",
            row.time_out or "",
            row.note or "",
        ]

        for c, value in enumerate(values, 1):
            cell = ws.cell(excel_row, c)
            cell.value = value
            cell.border = border
            cell.font = Font(name="Arial", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(
                "solid",
                fgColor=band_blue if idx % 2 == 1 else band_white,
            )

        ws.row_dimensions[excel_row].height = 24

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:F{max(len(rows) + 3, 4)}"


def generate_cbt_file(
    nova_path: str | Path | None,
    newstart_path: str | Path | None,
    hrn_path: str | Path | None,
    output_path: str | Path,
    work_date: date | None = None,
    newstart_tt_only: bool = True,
) -> dict:
    if nova_path is None and newstart_path is None and hrn_path is None:
        raise ValueError("No spreadsheets were provided. Upload at least one company spreadsheet.")

    nova_wb = load_workbook(nova_path, data_only=True) if nova_path is not None else None
    newstart_wb = load_workbook(newstart_path, data_only=True) if newstart_path is not None else None
    hrn_wb = load_workbook(hrn_path, data_only=True) if hrn_path is not None else None

    if work_date is None:
        workbook_candidates = [
            wb for wb in [nova_wb, newstart_wb, hrn_wb]
            if wb is not None
        ]

        found_date = None

        for wb in workbook_candidates:
            found_date = extract_date_from_workbook(wb)
            if found_date:
                break

        work_date = found_date or today_la()

    company_rows: list[tuple[str, list[AttendanceRow]]] = []
    counts: dict[str, int] = {}

    if nova_wb is not None:
        nova_rows = parse_nova(nova_wb)
        company_rows.append(("NOVA", nova_rows))
        counts["NOVA"] = len(nova_rows)

    if newstart_wb is not None:
        newstart_rows = parse_newstart(newstart_wb, tt_only=newstart_tt_only)
        company_rows.append(("Newstart", newstart_rows))
        counts["Newstart"] = len(newstart_rows)

    if hrn_wb is not None:
        hrn_rows = parse_hrn(hrn_wb, target_date=work_date)
        company_rows.append(("HRN", hrn_rows))
        counts["HRN"] = len(hrn_rows)

    out_wb = build_cbt_workbook(company_rows, work_date)
    out_wb.save(output_path)

    return {
        "output_path": str(output_path),
        "date": work_date.isoformat(),
        "counts": counts,
        "sheets": [
            f"{company}_{_mmdd(work_date)}"
            for company, _ in company_rows
        ],
    }
