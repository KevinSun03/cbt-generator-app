from __future__ import annotations

import base64
from datetime import date, datetime, time, timedelta
from io import BytesIO
import json
import re
from typing import Iterable, Sequence

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ============================================================================
# General helpers
# ============================================================================

def _clean(value) -> str:
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).strip(),
    )


def _name_key(name: str) -> str:
    return re.sub(
        r"[^a-z0-9\u4e00-\u9fff]+",
        "",
        _clean(name).lower(),
    )


def _image_data_url(uploaded_file) -> str:
    mime_type = (
        getattr(
            uploaded_file,
            "type",
            None,
        )
        or "image/jpeg"
    )

    encoded = base64.b64encode(
        uploaded_file.getvalue()
    ).decode("utf-8")

    return (
        f"data:{mime_type};base64,"
        f"{encoded}"
    )


def _combine_notes(*notes: str) -> str:
    unique_notes = []

    for note in notes:
        cleaned = _clean(note)

        if (
            cleaned
            and cleaned not in unique_notes
        ):
            unique_notes.append(cleaned)

    return "; ".join(unique_notes)


# ============================================================================
# Duplicate handling
# ============================================================================

def merge_duplicate_rows(
    rows: Iterable[dict],
) -> list[dict]:
    """
    Merge overlapping employee rows from multiple uploaded photos.

    Rows are matched by normalized employee name. If two photos contain
    conflicting times, the first value is retained and the conflict is
    added to OCR Notes for manual review.
    """
    merged: dict[str, dict] = {}
    order: list[str] = []

    confidence_rank = {
        "": 0,
        "low": 1,
        "medium": 2,
        "high": 3,
        "manual": 4,
    }

    for row in rows:
        name = _clean(row.get("Name"))
        key = _name_key(name)

        if not key:
            continue

        candidate = {
            "Name": name,
            "Check In": _clean(
                row.get("Check In")
            ),
            "Check Out": _clean(
                row.get("Check Out")
            ),
            "Confidence": (
                _clean(
                    row.get("Confidence")
                ).lower()
                or "low"
            ),
            "OCR Notes": _clean(
                row.get("OCR Notes")
            ),
        }

        if key not in merged:
            merged[key] = candidate
            order.append(key)
            continue

        current = merged[key]
        conflict_notes = []

        current_check_in = current["Check In"]
        candidate_check_in = (
            candidate["Check In"]
        )

        if (
            not current_check_in
            and candidate_check_in
        ):
            current["Check In"] = (
                candidate_check_in
            )

        elif (
            current_check_in
            and candidate_check_in
            and current_check_in
            != candidate_check_in
        ):
            conflict_notes.append(
                "Conflicting check-in values: "
                f"{current_check_in} / "
                f"{candidate_check_in}"
            )

        current_check_out = current["Check Out"]
        candidate_check_out = (
            candidate["Check Out"]
        )

        if (
            not current_check_out
            and candidate_check_out
        ):
            current["Check Out"] = (
                candidate_check_out
            )

        elif (
            current_check_out
            and candidate_check_out
            and current_check_out
            != candidate_check_out
        ):
            conflict_notes.append(
                "Conflicting check-out values: "
                f"{current_check_out} / "
                f"{candidate_check_out}"
            )

        current_confidence = (
            confidence_rank.get(
                current["Confidence"],
                0,
            )
        )

        candidate_confidence = (
            confidence_rank.get(
                candidate["Confidence"],
                0,
            )
        )

        if (
            candidate_confidence
            > current_confidence
        ):
            current["Confidence"] = (
                candidate["Confidence"]
            )

        current["OCR Notes"] = _combine_notes(
            current["OCR Notes"],
            candidate["OCR Notes"],
            *conflict_notes,
        )

    return [
        merged[key]
        for key in order
    ]


# ============================================================================
# OpenAI photo extraction
# ============================================================================

def extract_attendance_rows_from_images(
    image_files: Sequence,
    group_name: str,
    scheduled_in: time,
    scheduled_out: time,
    api_key: str,
    model: str = "gpt-5-mini",
) -> list[dict]:
    """
    Extract employee names and handwritten attendance times from one
    attendance group's uploaded photos.

    NOVA 1 and NOVA 2 are processed separately by the app, so each line
    can use a different schedule and different set of photos.
    """
    if not image_files:
        return []

    client = OpenAI(
        api_key=api_key
    )

    scheduled_in_text = (
        scheduled_in.strftime("%H:%M")
    )

    scheduled_out_text = (
        scheduled_out.strftime("%H:%M")
    )

    shift_crosses_midnight = (
        scheduled_out <= scheduled_in
    )

    overnight_instruction = (
        "This shift crosses midnight. Checkout occurs on the "
        "calendar day after check-in."
        if shift_crosses_midnight
        else
        "This shift does not cross midnight."
    )

    prompt = f"""
You are extracting employee attendance information from completed
attendance-sheet photos.

Attendance group: {group_name}
Scheduled check-in: {scheduled_in_text}
Scheduled check-out: {scheduled_out_text}

{overnight_instruction}

Read every uploaded image carefully.

The attendance form may contain:
- One table.
- Multiple pages.
- Two side-by-side employee tables.
- Printed names with handwritten check-in and checkout times.
- Blank attendance fields.
- Crossed-out rows.
- Repeated rows caused by overlapping photos.

Extraction rules:
1. Extract every real employee row that is not clearly crossed out.
2. Do not extract headings, dates, company names, totals, signatures,
   blank lines, or instructions as employee names.
3. Include an employee even when check-in or checkout is blank.
4. Read check-in only from the check-in column.
5. Read checkout only from the checkout column.
6. Return readable times in 24-hour HH:MM format.
7. Use the scheduled shift to infer whether handwriting means AM or PM.
8. Do not invent unreadable names or times.
9. If a value cannot be read, return an empty string and explain the
   uncertainty in OCR Notes.
10. Confidence must be exactly high, medium, or low.
11. Keep the original employee-name spelling visible in the photo.
""".strip()

    content = [
        {
            "type": "input_text",
            "text": prompt,
        }
    ]

    for image_file in image_files:
        content.append(
            {
                "type": "input_image",
                "image_url": (
                    _image_data_url(
                        image_file
                    )
                ),
                "detail": "high",
            }
        )

    schema = {
        "type": "object",
        "properties": {
            "rows": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "Name": {
                            "type": "string",
                        },
                        "Check In": {
                            "type": "string",
                        },
                        "Check Out": {
                            "type": "string",
                        },
                        "Confidence": {
                            "type": "string",
                            "enum": [
                                "high",
                                "medium",
                                "low",
                            ],
                        },
                        "OCR Notes": {
                            "type": "string",
                        },
                    },
                    "required": [
                        "Name",
                        "Check In",
                        "Check Out",
                        "Confidence",
                        "OCR Notes",
                    ],
                    "additionalProperties": False,
                },
            },
        },
        "required": [
            "rows",
        ],
        "additionalProperties": False,
    }

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "user",
                "content": content,
            }
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": (
                    "attendance_photo_rows"
                ),
                "strict": True,
                "schema": schema,
            }
        },
    )

    output_text = _clean(
        response.output_text
    )

    if not output_text:
        raise RuntimeError(
            "The vision model returned no attendance data."
        )

    try:
        payload = json.loads(
            output_text
        )
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            "The vision model returned invalid structured data."
        ) from exc

    extracted_rows = payload.get(
        "rows",
        [],
    )

    return merge_duplicate_rows(
        extracted_rows
    )


# ============================================================================
# Time parsing
# ============================================================================

def _clock_candidates(
    value,
) -> list[time]:
    """
    Parse entered time values.

    Supported examples:
    - 19:00
    - 1900
    - 7:00 PM
    - 7 PM
    - 7:00

    Ambiguous values such as 7:00 generate both 07:00 and 19:00.
    The candidate nearest to the scheduled time is selected later.
    """
    if value is None:
        return []

    if isinstance(value, datetime):
        return [
            value.time().replace(
                second=0,
                microsecond=0,
            )
        ]

    if isinstance(value, time):
        return [
            value.replace(
                second=0,
                microsecond=0,
            )
        ]

    text = _clean(value)

    if not text:
        return []

    text = (
        text
        .replace("：", ":")
        .replace("；", ":")
        .replace(";", ":")
        .replace(".", ":")
        .upper()
    )

    text = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    am_pm_match = re.fullmatch(
        r"(\d{1,2})"
        r"(?::(\d{1,2}))?"
        r"\s*(AM|PM)",
        text,
    )

    if am_pm_match:
        hour = int(
            am_pm_match.group(1)
        )

        minute = int(
            am_pm_match.group(2)
            or 0
        )

        am_pm = am_pm_match.group(3)

        if (
            hour < 1
            or hour > 12
            or minute > 59
        ):
            return []

        if am_pm == "AM":
            hour = (
                0
                if hour == 12
                else hour
            )
        else:
            hour = (
                12
                if hour == 12
                else hour + 12
            )

        return [
            time(hour, minute)
        ]

    compact_match = re.fullmatch(
        r"\d{3,4}",
        text,
    )

    if compact_match:
        digits = compact_match.group(0)
        hour = int(digits[:-2])
        minute = int(digits[-2:])

    else:
        regular_match = re.fullmatch(
            r"(\d{1,2})"
            r"(?::(\d{1,2}))?",
            text,
        )

        if not regular_match:
            return []

        hour = int(
            regular_match.group(1)
        )

        minute = int(
            regular_match.group(2)
            or 0
        )

    if minute > 59:
        return []

    if hour == 24 and minute == 0:
        return [
            time(0, 0)
        ]

    if hour < 0 or hour > 23:
        return []

    if 1 <= hour <= 11:
        return [
            time(hour, minute),
            time(hour + 12, minute),
        ]

    if hour == 12:
        return [
            time(12, minute),
            time(0, minute),
        ]

    return [
        time(hour, minute)
    ]


def _nearest_datetime_from_text(
    work_date: date,
    value,
    target: datetime,
) -> datetime | None:
    """
    Choose the interpreted date/time closest to the scheduled target.

    This allows:
    - Ambiguous AM/PM entries.
    - Overnight shifts.
    - Checkout times after midnight.
    """
    clocks = _clock_candidates(
        value
    )

    if not clocks:
        return None

    candidates = []

    for clock in clocks:
        for offset in (
            -1,
            0,
            1,
            2,
        ):
            candidate_date = (
                work_date
                + timedelta(days=offset)
            )

            candidates.append(
                datetime.combine(
                    candidate_date,
                    clock,
                )
            )

    return min(
        candidates,
        key=lambda candidate: abs(
            candidate - target
        ),
    )


def _minutes_text(
    minutes: int,
) -> str:
    if minutes == 1:
        return "1 minute"

    return f"{minutes} minutes"


# ============================================================================
# Attendance-rate calculation
# ============================================================================

def calculate_attendance_rate(
    group_name: str,
    records: Sequence[dict],
    work_date: date,
    scheduled_in: time,
    scheduled_out: time,
    expected_headcount: int,
) -> dict:
    """
    Calculate strict attendance.

    Qualified:
        actual check-in <= scheduled check-in
        AND
        actual checkout >= scheduled checkout

    One minute late or one minute early is an exception.
    """
    if expected_headcount <= 0:
        raise ValueError(
            "Expected headcount must be greater than zero."
        )

    scheduled_start = datetime.combine(
        work_date,
        scheduled_in,
    )

    scheduled_end = datetime.combine(
        work_date,
        scheduled_out,
    )

    if scheduled_end <= scheduled_start:
        scheduled_end += timedelta(
            days=1
        )

    normalized_records = (
        merge_duplicate_rows(
            records
        )
    )

    details: list[dict] = []
    observed_qualified = 0
    observed_exceptions = 0

    for record in normalized_records:
        name = _clean(
            record.get("Name")
        )

        if not name:
            continue

        check_in_text = _clean(
            record.get("Check In")
        )

        check_out_text = _clean(
            record.get("Check Out")
        )

        actual_in = (
            _nearest_datetime_from_text(
                work_date=work_date,
                value=check_in_text,
                target=scheduled_start,
            )
        )

        actual_out = (
            _nearest_datetime_from_text(
                work_date=work_date,
                value=check_out_text,
                target=scheduled_end,
            )
        )

        reasons = []

        if actual_in is None:
            reasons.append(
                "Missing or invalid check-in"
            )

        if actual_out is None:
            reasons.append(
                "Missing or invalid checkout"
            )

        if (
            actual_in is not None
            and actual_in > scheduled_start
        ):
            late_minutes = int(
                (
                    actual_in
                    - scheduled_start
                ).total_seconds()
                // 60
            )

            reasons.append(
                "Late by "
                + _minutes_text(
                    late_minutes
                )
            )

        if (
            actual_out is not None
            and actual_out < scheduled_end
        ):
            early_minutes = int(
                (
                    scheduled_end
                    - actual_out
                ).total_seconds()
                // 60
            )

            reasons.append(
                "Left early by "
                + _minutes_text(
                    early_minutes
                )
            )

        qualified = (
            actual_in is not None
            and actual_out is not None
            and actual_in
            <= scheduled_start
            and actual_out
            >= scheduled_end
        )

        if qualified:
            observed_qualified += 1
        else:
            observed_exceptions += 1

        details.append(
            {
                "Group": group_name,
                "Employee": name,
                "Scheduled In": (
                    scheduled_in.strftime(
                        "%H:%M"
                    )
                ),
                "Scheduled Out": (
                    scheduled_out.strftime(
                        "%H:%M"
                    )
                ),
                "Actual In": check_in_text,
                "Actual Out": check_out_text,
                "Status": (
                    "Qualified"
                    if qualified
                    else "Exception"
                ),
                "Exception": (
                    ""
                    if qualified
                    else "; ".join(
                        reasons
                    )
                ),
                "OCR Confidence": _clean(
                    record.get(
                        "Confidence"
                    )
                ),
                "OCR Notes": _clean(
                    record.get(
                        "OCR Notes"
                    )
                ),
            }
        )

    records_found = len(
        details
    )

    missing_records = max(
        expected_headcount
        - records_found,
        0,
    )

    extra_records = max(
        records_found
        - expected_headcount,
        0,
    )

    # Add unnamed placeholders for expected employees who have no record.
    for missing_index in range(
        1,
        missing_records + 1,
    ):
        details.append(
            {
                "Group": group_name,
                "Employee": (
                    "Missing scheduled employee "
                    f"#{missing_index}"
                ),
                "Scheduled In": (
                    scheduled_in.strftime(
                        "%H:%M"
                    )
                ),
                "Scheduled Out": (
                    scheduled_out.strftime(
                        "%H:%M"
                    )
                ),
                "Actual In": "",
                "Actual Out": "",
                "Status": "Exception",
                "Exception": (
                    "No attendance record found"
                ),
                "OCR Confidence": "",
                "OCR Notes": "",
            }
        )

    # Never allow the numerator to exceed the scheduled headcount.
    qualified_for_rate = min(
        observed_qualified,
        expected_headcount,
    )

    scheduled_exceptions = (
        expected_headcount
        - qualified_for_rate
    )

    attendance_rate = (
        qualified_for_rate
        / expected_headcount
        * 100
    )

    return {
        "summary": {
            "Group": group_name,
            "Scheduled Shift": (
                scheduled_in.strftime(
                    "%H:%M"
                )
                + "–"
                + scheduled_out.strftime(
                    "%H:%M"
                )
            ),
            "Expected": expected_headcount,
            "Records Found": records_found,
            "Missing Records": missing_records,
            "Extra Records": extra_records,
            "Observed Qualified": (
                observed_qualified
            ),
            "Observed Exceptions": (
                observed_exceptions
            ),
            "Qualified": (
                qualified_for_rate
            ),
            "Scheduled Exceptions": (
                scheduled_exceptions
            ),
            "Attendance Rate": (
                attendance_rate
            ),
        },
        "details": details,
    }


# ============================================================================
# Excel export
# ============================================================================

def build_attendance_rate_workbook(
    results: dict[str, dict],
    work_date: date,
) -> bytes:
    """
    Create an Excel workbook with:
    - Attendance_Summary
    - One detail sheet for each calculated group
    """
    if not results:
        raise ValueError(
            "No attendance results were provided."
        )

    workbook = Workbook()

    summary_ws = workbook.active
    summary_ws.title = (
        "Attendance_Summary"
    )

    dark_blue = "2F75B5"
    light_blue = "BDD7EE"
    light_row = "F3F6FC"
    alternate_row = "D9E8FA"
    qualified_green = "E2F0D9"
    exception_red = "FCE4D6"
    border_color = "A6A6A6"

    thin = Side(
        style="thin",
        color=border_color,
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    summary_headers = [
        "Group",
        "Scheduled Shift",
        "Expected",
        "Records Found",
        "Missing Records",
        "Extra Records",
        "Qualified",
        "Scheduled Exceptions",
        "Observed Exceptions",
        "Attendance Rate",
    ]

    summary_ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(
            summary_headers
        ),
    )

    summary_ws["A1"] = (
        "End-of-Day Attendance Rate — "
        f"{work_date.month}/"
        f"{work_date.day}/"
        f"{work_date.year}"
    )

    summary_ws["A1"].fill = (
        PatternFill(
            "solid",
            fgColor=dark_blue,
        )
    )

    summary_ws["A1"].font = Font(
        name="Arial",
        size=16,
        bold=True,
        color="FFFFFF",
    )

    summary_ws["A1"].alignment = (
        Alignment(
            horizontal="center",
            vertical="center",
        )
    )

    summary_ws.row_dimensions[
        1
    ].height = 32

    header_row = 3

    for column_number, header in enumerate(
        summary_headers,
        1,
    ):
        cell = summary_ws.cell(
            header_row,
            column_number,
            header,
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=light_blue,
        )

        cell.font = Font(
            name="Arial",
            bold=True,
            color="1F4E79",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = border

    for row_number, result in enumerate(
        results.values(),
        header_row + 1,
    ):
        summary = result["summary"]

        values = [
            summary["Group"],
            summary["Scheduled Shift"],
            summary["Expected"],
            summary["Records Found"],
            summary["Missing Records"],
            summary["Extra Records"],
            summary["Qualified"],
            summary[
                "Scheduled Exceptions"
            ],
            summary[
                "Observed Exceptions"
            ],
            summary[
                "Attendance Rate"
            ]
            / 100,
        ]

        row_fill = (
            alternate_row
            if row_number % 2 == 0
            else light_row
        )

        for column_number, value in enumerate(
            values,
            1,
        ):
            cell = summary_ws.cell(
                row_number,
                column_number,
                value,
            )

            cell.fill = PatternFill(
                "solid",
                fgColor=row_fill,
            )

            cell.border = border

            cell.font = Font(
                name="Arial",
                size=11,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        summary_ws.cell(
            row_number,
            10,
        ).number_format = "0.00%"

    summary_widths = [
        16,
        20,
        12,
        15,
        15,
        14,
        12,
        20,
        20,
        18,
    ]

    for column_number, width in enumerate(
        summary_widths,
        1,
    ):
        summary_ws.column_dimensions[
            get_column_letter(
                column_number
            )
        ].width = width

    summary_ws.freeze_panes = "A4"
    summary_ws.sheet_view.showGridLines = (
        False
    )

    summary_ws.auto_filter.ref = (
        f"A3:J"
        f"{max(summary_ws.max_row, 3)}"
    )

    summary_ws.page_setup.orientation = (
        "landscape"
    )

    summary_ws.page_setup.fitToWidth = 1
    summary_ws.page_setup.fitToHeight = 0
    summary_ws.sheet_properties.pageSetUpPr.fitToPage = (
        True
    )

    detail_headers = [
        "Group",
        "Employee",
        "Scheduled In",
        "Scheduled Out",
        "Actual In",
        "Actual Out",
        "Status",
        "Exception",
        "OCR Confidence",
        "OCR Notes",
    ]

    for group_name, result in (
        results.items()
    ):
        safe_group_name = re.sub(
            r"[\[\]:*?/\\]",
            "_",
            group_name,
        )[:20]

        worksheet = workbook.create_sheet(
            f"{safe_group_name}_Details"
        )

        for column_number, header in enumerate(
            detail_headers,
            1,
        ):
            cell = worksheet.cell(
                1,
                column_number,
                header,
            )

            cell.fill = PatternFill(
                "solid",
                fgColor=light_blue,
            )

            cell.font = Font(
                name="Arial",
                bold=True,
                color="1F4E79",
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            cell.border = border

        for row_number, detail in enumerate(
            result["details"],
            2,
        ):
            if (
                detail["Status"]
                == "Qualified"
            ):
                fill_color = (
                    qualified_green
                )
            else:
                fill_color = (
                    exception_red
                )

            for column_number, header in enumerate(
                detail_headers,
                1,
            ):
                cell = worksheet.cell(
                    row_number,
                    column_number,
                    detail.get(
                        header,
                        "",
                    ),
                )

                cell.fill = PatternFill(
                    "solid",
                    fgColor=fill_color,
                )

                cell.border = border

                cell.font = Font(
                    name="Arial",
                    size=11,
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

            worksheet.row_dimensions[
                row_number
            ].height = 30

        detail_widths = [
            14,
            30,
            15,
            15,
            15,
            15,
            14,
            38,
            18,
            40,
        ]

        for column_number, width in enumerate(
            detail_widths,
            1,
        ):
            worksheet.column_dimensions[
                get_column_letter(
                    column_number
                )
            ].width = width

        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = (
            False
        )

        worksheet.auto_filter.ref = (
            f"A1:J"
            f"{max(worksheet.max_row, 1)}"
        )

        worksheet.page_setup.orientation = (
            "landscape"
        )

        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = (
            True
        )

    output = BytesIO()
    workbook.save(output)

    return output.getvalue()
